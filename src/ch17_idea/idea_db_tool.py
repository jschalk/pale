from io import StringIO as io_StringIO
from numpy import float64
from openpyxl import load_workbook as openpyxl_load_workbook
from os import listdir as os_listdir
from os.path import (
    dirname as os_path_dirname,
    exists as os_path_exists,
    join as os_path_join,
)
from pandas import (
    DataFrame,
    ExcelWriter,
    read_csv as pandas_read_csv,
    read_excel as pandas_read_excel,
)
from pandas.api.types import is_numeric_dtype as pandas_api_types_is_numeric_dtype
from sqlite3 import Connection as sqlite3_Connection
from src.ch00_py.db_toolbox import (
    create_table_from_columns,
    create_table_from_csv,
    db_table_exists,
    get_table_columns,
    insert_csv,
)
from src.ch00_py.dict_toolbox import set_in_nested_dict
from src.ch00_py.file_toolbox import (
    create_path,
    get_dir_file_strs,
    get_dir_filenames,
    open_json,
    save_file,
    set_dir,
)
from src.ch16_translate.map_term import MapCore
from src.ch16_translate.translate_config import (
    get_translate_args_class_types,
    get_translateable_args,
)
from src.ch16_translate.translate_main import TranslateUnit, get_translateunit_from_dict
from src.ch17_idea._ref.ch17_semantic_types import FaceName, SparkInt
from src.ch17_idea.idea_config import (
    get_default_sorted_list,
    get_idea_dimen_ref,
    get_idea_elements_sort_order,
    get_idea_sqlite_types,
)


def save_dataframe_to_csv(x_df: DataFrame, x_dir: str, x_filename: str):
    save_file(x_dir, x_filename, get_ordered_csv(x_df))


def get_ordered_csv(x_df: DataFrame, sorting_columns: list[str] = None) -> str:
    new_sorting_columns = get_default_sorted_list(set(x_df.columns), sorting_columns)
    x_df = x_df.reindex(columns=new_sorting_columns)
    x_df.sort_values(new_sorting_columns, inplace=True)
    x_df.reset_index(inplace=True)
    x_df.drop(columns=["index"], inplace=True)
    return x_df.to_csv(index=False).replace("\r", "")


def open_csv(x_dir: str, x_filename: str = None) -> DataFrame:
    if os_path_exists(create_path(x_dir, x_filename)) is False:
        return None
    return pandas_read_csv(create_path(x_dir, x_filename))


def get_sheet_names(x_path: str) -> list[str]:
    return openpyxl_load_workbook(x_path).sheetnames


def get_all_excel_sheet_names(
    dir: str, sub_strs: set[str] = None
) -> set[(str, str, str)]:
    if sub_strs is None:
        sub_strs = set()
    excel_files = get_dir_filenames(dir, {"xlsx"})
    sheet_names = set()
    for relative_dir, filename in excel_files:
        absolute_dir = create_path(dir, relative_dir)
        absolute_path = create_path(absolute_dir, filename)
        file_sheet_names = get_sheet_names(absolute_path)
        for sheet_name in file_sheet_names:
            if not sub_strs:
                sheet_names.add((absolute_dir, filename, sheet_name))
            else:
                for sub_str in sub_strs:
                    if sheet_name.find(sub_str) >= 0:
                        sheet_names.add((absolute_dir, filename, sheet_name))
    return sheet_names


def get_relevant_columns_dataframe(
    src_df: DataFrame, relevant_columns: list[str] = None
) -> DataFrame:
    if relevant_columns is None:
        relevant_columns = get_idea_elements_sort_order()
    current_columns = set(src_df.columns.to_list())
    relevant_columns_set = set(relevant_columns)
    current_relevant_columns = current_columns & (relevant_columns_set)
    relevant_cols_in_order = [
        r_col for r_col in relevant_columns if r_col in current_relevant_columns
    ]
    return src_df[relevant_cols_in_order]


def get_dataframe_translateable_columns(x_df: DataFrame) -> set[str]:
    return {
        x_column for x_column in x_df.columns if x_column in get_translateable_args()
    }


def translate_single_column_dataframe(
    x_df: DataFrame, x_mapunit: MapCore, column_name: str
) -> DataFrame:
    if column_name in x_df:
        row_count = len(x_df)
        for cur_row in range(row_count):
            otx_value = x_df.iloc[cur_row][column_name]
            inx_value = x_mapunit.reveal_inx(otx_value)
            x_df.at[cur_row, column_name] = inx_value
    return x_df


def translate_all_columns_dataframe(x_df: DataFrame, x_translateunit: TranslateUnit):
    if x_translateunit is None:
        return None

    column_names = set(x_df.columns)
    translateable_columns = column_names & (get_translateable_args())
    for translateable_column in translateable_columns:
        class_type = get_translate_args_class_types().get(translateable_column)
        x_mapunit = x_translateunit.get_mapunit(class_type)
        translate_single_column_dataframe(x_df, x_mapunit, translateable_column)


def move_otx_csvs_to_translate_inx(face_dir: str):
    otx_dir = create_path(face_dir, "otx")
    inx_dir = create_path(face_dir, "inx")
    translate_filename = "translate.json"
    translate_dict = open_json(face_dir, translate_filename)
    face_translateunit = get_translateunit_from_dict(translate_dict)
    otx_dir_files = get_dir_file_strs(otx_dir, delete_extensions=False)
    for x_filename in otx_dir_files.keys():
        x_df = open_csv(otx_dir, x_filename)
        translate_all_columns_dataframe(x_df, face_translateunit)
        save_dataframe_to_csv(x_df, inx_dir, x_filename)


def _get_translate_idea_format_filenames() -> set[str]:
    idea_numbers = set(get_idea_dimen_ref().get("translate_name"))
    idea_numbers.update(set(get_idea_dimen_ref().get("translate_title")))
    idea_numbers.update(set(get_idea_dimen_ref().get("translate_label")))
    idea_numbers.update(set(get_idea_dimen_ref().get("translate_rope")))
    return {f"{idea_number}.xlsx" for idea_number in idea_numbers}


def append_df_to_excel(file_path: str, sheet_name: str, dataframe: DataFrame):
    try:
        # Load the existing workbook
        workbook = openpyxl_load_workbook(file_path)

        # Check if the sheet exists, if not create it
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            # Add column names to the new sheet
            for col_num, column_header in enumerate(dataframe.columns, 1):
                sheet.cell(row=1, column=col_num, value=column_header)
            start_row = 2  # Start appending data from the second row
        else:
            sheet = workbook[sheet_name]
            start_row = sheet.max_row + 1

        # Convert the DataFrame to a list of rows
        rows = dataframe.to_dict(orient="split")["data"]

        # Append the rows to the sheet
        for i, row in enumerate(rows, start_row):
            for j, value in enumerate(row, 1):  # 1-based index for Excel
                sheet.cell(row=i, column=j, value=value)

        # Save changes to the workbook
        workbook.save(file_path)
        # prt("Data appended successfully!")

    except FileNotFoundError:
        # If the file doesn't exist, create a new one
        # prt(f"{file_path} not found. Creating a new Excel file.")
        dataframe.to_excel(file_path, index=False, sheet_name=sheet_name)


class pandas_tools_ExcelWriterException(Exception):
    pass


def upsert_sheet(
    file_path: str, sheet_name: str, dataframe: DataFrame, replace: bool = False
):
    # sourcery skip: remove-redundant-exception, simplify-single-exception-tuple
    set_dir(os_path_dirname(file_path))
    """
    Updates or creates an Excel sheet with a specified DataFrame.

    Args:
    - file_path (str): The path to the Excel file.
    - sheet_name (str): The name of the sheet to update or create.
    - dataframe (pd.DataFrame): The DataFrame to write to the sheet.
    """
    # Check if the file exists
    if not os_path_exists(file_path):
        # If file does not exist, create it with the specified sheet
        with ExcelWriter(file_path, engine="xlsxwriter") as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # If the file exists, check for the sheet
    try:
        if replace:
            with ExcelWriter(
                file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            append_df_to_excel(file_path, sheet_name, dataframe)

    except (PermissionError, FileNotFoundError, OSError) as e:
        raise pandas_tools_ExcelWriterException(f"An error occurred: {e}") from e


def sheet_exists(file_path: str, sheet_name: str):
    """
    Checks if a specific sheet exists in an Excel workbook.

    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to check.

    Returns:
        bool: True if the sheet exists, False otherwise.
    """
    if not os_path_exists(file_path):
        return False

    try:
        return sheet_name in set(get_sheet_names(file_path))
    except Exception as e:
        return False


def split_excel_into_dirs(
    src_file: str, dst_dir: str, column_name: str, filename: str, sheet_name: str
):
    """
    Splits an Excel file into multiple Excel files, each containing rows
    corresponding to a unique value in the specified column.

    Args:
        src_file (str): Path to the src Excel file.
        dst_dir (str): Directory where the files will be saved.
        column_name (str): Column to split by unique values.
    """
    # Create the destination directory if it doesn't exist
    set_dir(dst_dir)
    df = pandas_read_excel(src_file, sheet_name=sheet_name)

    # Check if the column exists
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' does not exist in the src file.")

    # Group by unique values in the column
    unique_values = df[column_name].unique()

    for value in unique_values:
        if float64 != type(value):
            # Filter rows for the current unique value
            filtered_df = df[df[column_name] == value]

            # Create a safe subdirectory name for the unique value
            safe_value = str(value).replace("/", "_").replace("\\", "_")
            subdirectory = create_path(dst_dir, safe_value)
            # Create the subdirectory if it doesn't exist
            set_dir(subdirectory)
            # Define the destination file path
            dst_file = create_path(subdirectory, f"{filename}.xlsx")
            upsert_sheet(dst_file, sheet_name, filtered_df)


def if_nan_return_None(x_obj: any) -> any:
    # sourcery skip: equality-identity, remove-redundant-if
    # If the value is NaN, the comparison value != value
    return None if x_obj != x_obj else x_obj


def dataframe_to_dict(x_df: DataFrame, key_columns: list[str]) -> dict:
    df_dict = x_df.to_dict(orient="records")
    x_dict = {}
    for x_value in df_dict:
        if x_value.get("id"):
            x_value.pop("id")
        nested_keys = [x_value[key_column] for key_column in key_columns]
        set_in_nested_dict(x_dict, nested_keys, x_value)
    return x_dict


def create_idea_table_from_csv(
    csv_filepath: str, conn_or_cursor: sqlite3_Connection, tablename: str
):
    column_types = get_idea_sqlite_types()
    create_table_from_csv(csv_filepath, conn_or_cursor, tablename, column_types)


def insert_idea_csv(
    csv_filepath: str, conn_or_cursor: sqlite3_Connection, tablename: str
):
    if db_table_exists(conn_or_cursor, tablename) is False:
        create_idea_table_from_csv(csv_filepath, conn_or_cursor, tablename)

    # Future feature? filtering csv file so only relevant idea columns are loaded
    insert_csv(csv_filepath, conn_or_cursor, tablename)


def get_pragma_table_fetchall(table_columns):
    pragma_table_attrs = []
    idea_sqlite_types = get_idea_sqlite_types()
    for x_count, x_col in enumerate(table_columns):
        col_type = idea_sqlite_types.get(x_col)
        pragma_table_attrs.append((x_count, x_col, col_type, 0, None, 0))
    return pragma_table_attrs


def save_table_to_csv(conn_or_cursor: sqlite3_Connection, dst_dir: str, tablename: str):
    """given a cursor object, a directory, a tablename create tablename.csv file"""

    select_sqlstr = f"""SELECT * FROM {tablename};"""
    tables_rows = conn_or_cursor.execute(select_sqlstr).fetchall()
    tables_columns = get_table_columns(conn_or_cursor, tablename)
    # momentunit_columns = [desc[0] for desc in cursor.description]
    table_df = DataFrame(tables_rows, columns=tables_columns)
    table_filename = f"{tablename}.csv"
    save_dataframe_to_csv(table_df, dst_dir, table_filename)


def create_idea_sorted_table(
    conn: sqlite3_Connection, tablename: str, columns_list: list[str]
):
    columns_list = get_default_sorted_list(set(columns_list))
    create_table_from_columns(conn, tablename, columns_list, get_idea_sqlite_types())


def get_idea_into_dimen_raw_query(
    conn_or_cursor: sqlite3_Connection,
    idea_number: str,
    x_dimen: str,
    x_jkeys: set[str],
    action_str: str = None,
) -> str:
    src_table = f"{idea_number}_raw"
    src_columns = get_table_columns(conn_or_cursor, src_table)
    dst_table = f"{x_dimen}_put_raw" if action_str else f"{x_dimen}_raw"
    dst_columns = get_table_columns(conn_or_cursor, dst_table)
    common_columns_set = set(dst_columns) & (set(src_columns))
    common_columns_list = [col for col in dst_columns if col in common_columns_set]
    common_columns_header = ", ".join(common_columns_list)
    values_cols = set(common_columns_set)
    values_cols.difference_update(x_jkeys)
    return f"""INSERT INTO {dst_table} (idea_number, {common_columns_header})
SELECT '{idea_number}' as idea_number, {common_columns_header}
FROM {src_table}
{_get_keys_where_str(x_jkeys, dst_columns)}
GROUP BY {common_columns_header}
;
"""


def _get_keys_where_str(x_jkeys: set[str], dst_columns: list[str]) -> str:
    key_columns_list = [col for col in dst_columns if col in x_jkeys]
    keys_where_str = None
    for x_jkey in key_columns_list:
        if keys_where_str is None:
            keys_where_str = f"WHERE {x_jkey} IS NOT NULL"
        else:
            keys_where_str += f" AND {x_jkey} IS NOT NULL"
    return "" if keys_where_str is None else keys_where_str


def csv_dict_to_excel(csv_dict: dict[str, str], dir: str, filename: str):
    """
    Converts a dictionary of CSV strings into an Excel file.

    :param csv_dict: Dictionary where keys are sheet names and values are CSV strings
    :param file_path: Path to save the Excel file
    """
    set_dir(dir)
    file_path = create_path(dir, filename)
    x_excelwriter = ExcelWriter(file_path, engine="xlsxwriter")

    for sheet_name, csv_str in csv_dict.items():
        df = pandas_read_csv(io_StringIO(csv_str))  # Convert CSV string to DataFrame
        # Excel sheet names max length is 31 chars
        df.to_excel(x_excelwriter, sheet_name=sheet_name[:31], index=False)

    x_excelwriter.close()


def set_dataframe_first_two_columns(df: DataFrame, value_col1, value_col2) -> DataFrame:
    """
    Sets the first and second columns of a pandas DataFrame to specified values.

    Parameters:
    - df (pd.DataFrame): The DataFrame to modify.
    - value_col1: The value to set in the first column.
    - value_col2: The value to set in the second column.

    Returns:
    - pd.DataFrame: The modified DataFrame.
    """
    if df.shape[1] < 2:
        raise ValueError("DataFrame must have at least two columns.")

    df.iloc[:, 0] = value_col1
    df.iloc[:, 1] = value_col2
    return df


def check_dataframe_column_names(df: DataFrame, name_col1: str, name_col2: str) -> bool:
    """
    Checks if the first two columns of a pandas DataFrame have the specified names.

    Parameters:
    - df (pd.DataFrame): The DataFrame to check.
    - name_col1 (str): Expected name of the first column.
    - name_col2 (str): Expected name of the second column.

    Returns:
    - bool: True if the first two columns have the correct names, False otherwise.
    """
    if df.shape[1] < 2:
        raise ValueError("DataFrame must have at least two columns.")

    return df.columns[0] == name_col1 and df.columns[1] == name_col2


def update_all_face_name_spark_num_columns(
    excel_file_path: str, face_name: FaceName, spark_num: SparkInt
):
    workbook = openpyxl_load_workbook(excel_file_path)
    # Loop through all sheets in the workbook
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        if ws["A1"].value == "spark_num" and ws["B1"].value == "face_name":
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=1, value=spark_num)
                ws.cell(row=row, column=2, value=face_name)

            # Save the updated sheet
            workbook.save(excel_file_path)


class sqlite_data_type_Exception(Exception):
    pass


def is_column_type_valid(df: DataFrame, column: str, sqlite_data_type: str) -> bool:
    """expected sqlite_data_types: INT, REAL, TEXT"""
    if sqlite_data_type == "INT":
        expected_data_type = "int64"
    elif sqlite_data_type == "REAL":
        expected_data_type = "float64"
    elif sqlite_data_type == "TEXT":
        expected_data_type = "object"
    else:
        raise sqlite_data_type_Exception(f"{sqlite_data_type} is not valid sqlite_type")
    if column not in df.columns:
        return False
    # If column is completely empty (all NaN), accept it
    if df[column].isna().all():
        return True
    actual_dtype = df[column].dropna().infer_objects().dtype
    return str(actual_dtype) == expected_data_type


def prettify_excel(input_file: str, zoom: int = 120) -> None:
    """
    Reads all sheets from an Excel file, applies formatting improvements to each,
    and overwrites the original file. Safely handles sheets with only headers and no data.

    Args:
        input_file (str): Path to the Excel file to overwrite.
        zoom (int): Zoom level for each worksheet.
    """
    # Load all sheets
    sheet_data = pandas_read_excel(input_file, sheet_name=None)

    with ExcelWriter(input_file, engine="xlsxwriter") as writer:
        for sheet_name, df in sheet_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            worksheet.freeze_panes(1, 0)
            worksheet.set_zoom(zoom)

            # Format header
            header_format = workbook.add_format(
                {
                    "bold": True,
                    "text_wrap": True,
                    "valign": "top",
                    "fg_color": "#D7E4BC",
                    "border": 1,
                }
            )
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)

            # Format columns and set widths
            for i, col in enumerate(df.columns):
                col_data = df[col]
                width = (
                    max(
                        (0 if col_data.empty else col_data.astype(str).map(len).max()),
                        len(str(col)),
                    )
                    + 2
                )

                if pandas_api_types_is_numeric_dtype(col_data):
                    if "salary" in col.lower() or "amount" in col.lower():
                        fmt = workbook.add_format({"num_format": "$#,##0", "border": 1})
                    else:
                        fmt = workbook.add_format({"num_format": "0.00", "border": 1})
                else:
                    fmt = workbook.add_format({"border": 1})
                worksheet.set_column(i, i, width, fmt)
            # Add table only if DataFrame has at least one row
            if not df.empty:
                worksheet.add_table(
                    0,
                    0,
                    len(df),
                    len(df.columns) - 1,
                    {
                        "columns": [{"header": col} for col in df.columns],
                        "style": "Table Style Medium 9",
                    },
                )


def update_spark_num_in_excel_files(directory: str, value) -> None:
    """
    Adds or updates the 'spark_num' column with a given value
    in all Excel files in the directory that contain 'stance' in the filename.

    Args:
        directory (str): Path to the directory containing Excel files.
        value: The value to set in the 'spark_num' column.
    """
    for filename in os_listdir(directory):
        if (
            filename.lower().endswith((".xlsx", ".xls"))
            and "stance" in filename.lower()
        ):
            filepath = os_path_join(directory, filename)

            # Read all sheets
            sheets = pandas_read_excel(filepath, sheet_name=None)

            # Modify each sheet
            updated_sheets = {}
            for sheet_name, df in sheets.items():
                df["spark_num"] = value  # Add or overwrite
                updated_sheets[sheet_name] = df

            # Write all sheets back to the same file
            with ExcelWriter(filepath, engine="xlsxwriter") as writer:
                for sheet_name, df in updated_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
