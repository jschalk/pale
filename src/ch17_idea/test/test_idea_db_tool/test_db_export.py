from openpyxl import load_workbook
import pytest
import sqlite3
from src.ch17_idea.idea_db_tool import export_db_to_excel


@pytest.fixture
def db():
    conn = sqlite3.connect(":memory:")
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE users (id INTEGER, name TEXT, age INTEGER)")
    cursor.executemany(
        "INSERT INTO users VALUES (?, ?, ?)",
        [(1, "Alice", 30), (2, "Bob", 25), (3, "Charlie", 35)],
    )
    cursor.execute("CREATE TABLE products (id INTEGER, title TEXT, price REAL)")
    cursor.executemany(
        "INSERT INTO products VALUES (?, ?, ?)",
        [(1, "Widget", 9.99), (2, "Gadget", 24.99)],
    )
    conn.commit()
    yield cursor
    conn.close()


def test_export_db_to_excel_CreatesFile_Exists(db, tmp_path):
    # ESTABLISH
    dest = str(tmp_path / "output.xlsx")
    # WHEN
    export_db_to_excel(db, dest)
    # THEN
    assert (tmp_path / "output.xlsx").exists()


def test_export_db_to_excel_CreatesFile_sheet_per_table(db, tmp_path):
    # ESTABLISH
    dest = str(tmp_path / "output.xlsx")
    # WHEN
    export_db_to_excel(db, dest)
    # THEN
    wb = load_workbook(dest)
    assert set(wb.sheetnames) == {"users", "products"}


def test_export_db_to_excel_CreatesFile_headers(db, tmp_path):
    # ESTABLISH
    dest = str(tmp_path / "output.xlsx")
    # WHEN
    export_db_to_excel(db, dest)
    # THEN
    wb = load_workbook(dest)
    headers = [cell.value for cell in wb["users"][1]]
    assert headers == ["id", "name", "age"]


def test_export_db_to_excel_CreatesFile_row_count(db, tmp_path):
    # ESTABLISH
    dest = str(tmp_path / "output.xlsx")
    # WHEN
    export_db_to_excel(db, dest)
    # THEN
    wb = load_workbook(dest)
    # max_row includes header row
    assert wb["users"].max_row == 4
    assert wb["products"].max_row == 3


def test_export_db_to_excel_CreatesFile_data_values(db, tmp_path):
    # ESTABLISH
    dest = str(tmp_path / "output.xlsx")
    # WHEN
    export_db_to_excel(db, dest)
    wb = load_workbook(dest)
    # THEN
    rows = [[cell.value for cell in row] for row in wb["users"].iter_rows(min_row=2)]
    assert rows == [[1, "Alice", 30], [2, "Bob", 25], [3, "Charlie", 35]]


def test_export_db_to_excel_CreatesFile_empty_table(db, tmp_path):
    # ESTABLISH
    db.execute("CREATE TABLE empty_table (id INTEGER, value TEXT)")
    dest = str(tmp_path / "output.xlsx")
    # WHEN
    export_db_to_excel(db, dest)
    # THEN
    wb = load_workbook(dest)
    assert "empty_table" in wb.sheetnames
    assert wb["empty_table"].max_row == 1  # header only


def test_export_db_to_excel_CreatesFile_no_tables_raises(tmp_path):
    # ESTABLISH
    conn = sqlite3.connect(":memory:")
    cursor = conn.cursor()
    # WHEN / THEN
    with pytest.raises(ValueError, match="No tables found"):
        export_db_to_excel(cursor, str(tmp_path / "output.xlsx"))
    conn.close()
