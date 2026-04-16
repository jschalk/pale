from openpyxl import load_workbook
from os.path import exists as os_path_exists, join as os_path_join
from pandas import (
    DataFrame,
    ExcelWriter as pandas_ExcelWriter,
    read_excel as pandas_read_excel,
)
from pandas.testing import assert_frame_equal as pandas_testing_assert_frame_equal
import pytest
from src.ch00_py.file_toolbox import create_path
from src.ch17_idea.idea_db_tool import (
    csv_dict_to_excel,
    get_idea_sqlite_types,
    prettify_excel_file,
    prettify_excel_files,
    set_df_idea_column_types,
)
from src.ref.keywords import Ch17Keywords as kw
from unittest.mock import MagicMock, patch


def test_csv_dict_to_excel_SavesFile(temp3_fs):
    # ESTABLISH
    test_data = {"TestSheet": "A,B\n5,6\n7,8"}
    x_dir = str(temp3_fs)
    x_filename = "test_data.xlsx"
    file_path = create_path(x_dir, x_filename)
    assert os_path_exists(file_path) is False

    # WHEN
    csv_dict_to_excel(test_data, x_dir, x_filename)

    # THEN
    assert os_path_exists(file_path)
    # Load the created Excel file to verify its contents
    df = pandas_read_excel(file_path, sheet_name="TestSheet")
    expected_df = DataFrame({"A": [5, 7], "B": [6, 8]})

    pandas_testing_assert_frame_equal(df, expected_df)
    print("Test passed successfully.")


def test_prettify_excel_file_SetsAttrs(temp3_fs):
    # sourcery skip: no-loop-in-tests, no-conditionals-in-tests
    # ESTABLISH
    temp_dir = str(temp3_fs)
    file_path = os_path_join(temp_dir, "test_huh.xlsx")

    df1 = DataFrame({"Name": ["Alice", "Bob"], "Salary": [50000, 60000]})
    df2 = DataFrame({"Item": ["Pen", "Notebook"], "Price": [1.5, 3.0]})

    with pandas_ExcelWriter(file_path, engine="xlsxwriter") as writer:
        df1.to_excel(writer, sheet_name="Employees", index=False)
        df2.to_excel(writer, sheet_name="Supplies", index=False)

    # WHEN
    prettify_excel_file(file_path)

    # THEN: Verify formatting changes
    wb = load_workbook(file_path)

    # Check headers in both sheets
    for sheet_name in ["Employees", "Supplies"]:
        ws = wb[sheet_name]

        # # Confirm header fill color (should match '#D7E4BC' -> RGB: D7E4BC)
        header_fill = ws["A1"].fill
        print(f"{header_fill=}")
        # assert isinstance(header_fill, PatternFill)
        # assert header_fill.fill_type == "solid"
        # assert header_fill.start_color.rgb in (
        #     "FFD7E4BC",
        #     "00D7E4BC",
        # )  # OpenPyXL sometimes uses different alpha

        # Confirm freeze pane is set at row 2
        print(f"{ws.freeze_panes=}")
        assert ws.freeze_panes == "A2"

        # Confirm zoom level (optional: may not always be persisted depending on openpyxl)
        if ws.sheet_view.zoomScale is not None:
            assert ws.sheet_view.zoomScale == 120

        # Confirm at least one column has adjusted width (Excel doesn't save actual width in standard units)
        col_widths = [
            ws.column_dimensions[col_letter].width for col_letter in ["A", "B"]
        ]
        assert any(
            width and width > 8 for width in col_widths
        )  # default width is ~8.43


def test_prettify_excel_files_SetsAttrs(temp3_fs):
    # sourcery skip: no-loop-in-tests, no-conditionals-in-tests
    # ESTABLISH
    temp_dir = str(temp3_fs)
    file_path = os_path_join(temp_dir, "test_huh.xlsx")

    df1 = DataFrame({"Name": ["Alice", "Bob"], "Salary": [50000, 60000]})
    df2 = DataFrame({"Item": ["Pen", "Notebook"], "Price": [1.5, 3.0]})

    with pandas_ExcelWriter(file_path, engine="xlsxwriter") as writer:
        df1.to_excel(writer, sheet_name="Employees", index=False)
        df2.to_excel(writer, sheet_name="Supplies", index=False)
    print(f"{temp_dir=}")
    print(f"{type(temp_dir)=}")

    # WHEN
    prettify_excel_files(temp_dir)

    # THEN: Verify formatting changes
    wb = load_workbook(file_path)

    # Check headers in both sheets
    for sheet_name in ["Employees", "Supplies"]:
        ws = wb[sheet_name]

        # # Confirm header fill color (should match '#D7E4BC' -> RGB: D7E4BC)
        header_fill = ws["A1"].fill
        print(f"{header_fill=}")
        # assert isinstance(header_fill, PatternFill)
        # assert header_fill.fill_type == "solid"
        # assert header_fill.start_color.rgb in (
        #     "FFD7E4BC",
        #     "00D7E4BC",
        # )  # OpenPyXL sometimes uses different alpha

        # Confirm freeze pane is set at row 2
        print(f"{ws.freeze_panes=}")
        assert ws.freeze_panes == "A2"

        # Confirm zoom level (optional: may not always be persisted depending on openpyxl)
        if ws.sheet_view.zoomScale is not None:
            assert ws.sheet_view.zoomScale == 120

        # Confirm at least one column has adjusted width (Excel doesn't save actual width in standard units)
        col_widths = [
            ws.column_dimensions[col_letter].width for col_letter in ["A", "B"]
        ]
        assert any(
            width and width > 8 for width in col_widths
        )  # default width is ~8.43


def test_set_df_idea_column_types_SetsAttrs_Scenario0_BasicConversion():
    # ESTABLISH
    df = DataFrame(
        {
            kw.plan_label: ["a", "b"],
            kw.addin: ["1", "2"],
            kw.gogo_want: ["10.5", "20.1"],
        }
    )
    # WHEN
    result = set_df_idea_column_types(df)
    # THEN
    assert str(result[kw.plan_label].dtype) == "str"
    assert str(result[kw.addin].dtype) == "int64"
    assert str(result[kw.gogo_want].dtype) == "float64"
    assert result[kw.addin].tolist() == [1, 2]
    assert result[kw.gogo_want].tolist() == [10.5, 20.1]


def test_set_df_idea_column_types_SetsAttrs_Scenario1_HandlesInvalidValuesWithCoerce():
    # ESTABLISH
    df = DataFrame({kw.numor: ["1", "bad", "3"]})
    # WHEN
    result = set_df_idea_column_types(df)
    # THEN
    assert str(result[kw.numor].dtype) == "Int64"
    assert result[kw.numor].isna().tolist() == [False, True, False]


def test_set_df_idea_column_types_SetsAttrs_Scenario2_MissingColumnIsIgnored():
    # ESTABLISH
    df = DataFrame({kw.numor: ["1", "2"]})
    # WHEN
    result = set_df_idea_column_types(df)
    # THEN
    assert "b" not in result.columns
    assert str(result[kw.numor].dtype) == "Int64"


# def test_set_df_idea_column_types_SetsAttrs_Scenario3_Unsupported_dtype_raises_Exception():
#     # ESTABLISH
#     df = DataFrame({kw.plan_label: ["1", "2"]})
#     # WHEN/THEN
#     with pytest.raises(ValueError, match="Unsupported dtype"):
#         set_df_idea_column_types(df)


def test_set_df_idea_column_types_SetsAttrs_Scenario4_DoesNotMutateOriginalDataframe():
    # ESTABLISH
    df = DataFrame({kw.numor: ["1", "2"]})
    # WHEN
    result = set_df_idea_column_types(df)
    # THEN
    # original should remain object/string-like
    assert str(df[kw.numor].dtype) in {"object", "str"}
    assert str(result[kw.numor].dtype) == "Int64"
