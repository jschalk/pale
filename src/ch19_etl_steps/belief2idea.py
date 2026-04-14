from openpyxl import load_workbook
from os import listdir as os_listdir
from os.path import join as os_path_join
from pandas import (
    DataFrame,
    ExcelWriter,
    read_excel as pandas_read_excel,
    to_numeric as pandas_to_numeric,
)
from pathlib import Path
from src.ch00_py.dict_toolbox import get_0_if_None
from src.ch00_py.file_toolbox import delete_dir, set_dir
from src.ch17_idea.idea_config import get_idea_types
from src.ch17_idea.idea_db_tool import save_sheet
from typing import List, Tuple


def get_spark_faces_from_df(df: DataFrame) -> set:
    """
    Returns a set of distinct values from the 'spark_face' column.
    NaN values are excluded.
    If the column does not exist, returns an empty set.
    """
    if "spark_face" not in df.columns:
        return set()

    return set(df["spark_face"].dropna().unique().tolist())


def get_spark_faces_from_files(directory) -> set:
    """
    Given a directory, read all Excel files and return a set of all distinct
    spark_face values across all sheets in all files.

    Uses get_spark_faces_from_df for per-sheet extraction.
    """
    all_faces = set()
    directory = Path(directory)

    for file_path in directory.iterdir():
        if not file_path.is_file():
            continue

        if file_path.suffix.lower() not in {".xlsx", ".xls"}:
            continue

        # Read all sheets
        sheets = pandas_read_excel(file_path, sheet_name=None)

        for df in sheets.values():
            faces = get_spark_faces_from_df(df)
            all_faces.update(faces)

    return all_faces


def get_max_spark_num_from_files(directory) -> int | None:
    """
    Returns the maximum integer spark_num across all Excel files and sheets.

    - Ignores missing, empty, and non-numeric values
    - Converts floats to ints
    - Returns None if no valid spark_num is found
    """
    directory = Path(directory)
    max_val = None

    for file_path in directory.iterdir():
        if not file_path.is_file():
            continue
        if file_path.suffix.lower() not in {".xlsx", ".xls"}:
            continue

        sheets = pandas_read_excel(file_path, sheet_name=None)
        for df in sheets.values():
            max_val = get_max_spark_num_from_df(df, max_val)
    return max_val


def get_max_spark_num_from_df(df: DataFrame, max_val: int) -> int:
    if "spark_num" not in df.columns:
        return max_val

    # Convert to numeric, coerce errors to NaN
    numeric_series = pandas_to_numeric(df["spark_num"], errors="coerce").dropna()

    if numeric_series.empty:
        return max_val

    # Convert floats to ints
    numeric_series = numeric_series.astype(int)

    current_max = numeric_series.max()

    if max_val is None or current_max > max_val:
        max_val = int(current_max)
    return max_val


def create_spark_face_spark_nums(
    spark_faces: set[str], max_spark_num: int = None
) -> dict[str, int]:
    if max_spark_num is None:
        max_spark_num = 0
    return {
        spark_face: max_spark_num + x_count
        for x_count, spark_face in enumerate(sorted(list(spark_faces)), start=1)
    }


def add_spark_num_column(df: DataFrame, spark_face_spark_nums: dict[str, int]):
    """
    Adds 'spark_num' as the first column based on 'spark_face' values.
    - mutates original DataFrame (does not create new df)
    """
    if "spark_num" in df.columns:
        df.drop(columns=["spark_num"], inplace=True)

    if "spark_face" not in df.columns:
        # raise ValueError("Column 'spark_face' not found in DataFrame")
        return
    spark_num_series = df["spark_face"].map(spark_face_spark_nums)

    # Insert as first column
    df.insert(0, "spark_num", spark_num_series)


def get_excel_sheet_tuples(directory: str) -> List[Tuple[str, str]]:
    """
    Given a directory, returns a sorted list of (filename, sheet_name) tuples
    for all Excel files found in that directory.

    Args:
        directory: Path to the directory to search for Excel files.

    Returns:
        Sorted list of (filename, sheet_name) tuples.
    """
    result = []
    excel_extensions = (".xlsx", ".xlsm", ".xltx", ".xltm")

    for filename in os_listdir(directory):
        if filename.lower().endswith(excel_extensions):
            filepath = os_path_join(directory, filename)
            wb = load_workbook(filepath, read_only=True)
            result.extend((filename, sheet_name) for sheet_name in wb.sheetnames)
            wb.close()

    return sorted(result)


def get_sheets_with_idea_types(directory: str) -> List[Tuple[str, str]]:
    """
    Returns all (filename, sheet_name) tuples where the sheet_name contains
    any of the provided idea_types.

    Args:
        directory:  Path to the directory to search for Excel files.
        idea_types: Set of strings to match against sheet names.

    Returns:
        Sorted list of (filename, sheet_name) tuples where sheet_name
        contains at least one idea_type.
    """
    idea_types = get_idea_types()
    all_tuples = get_excel_sheet_tuples(directory)
    return [
        (filename, sheet_name)
        for filename, sheet_name in all_tuples
        if any(idea_type in sheet_name.lower() for idea_type in idea_types)
    ]


def get_validated_b_src_idea_type_sheets(
    b_src_dir: str,
    i_src_dir: str,
) -> List[Tuple[str, str]]:
    """
    Returns all idea_type sheets found in b_src_dir.
    Raises a ValueError if any of those idea_type sheets also exist in i_src_dir.

    Args:
        b_src_dir: Path to the BELE source directory.
        i_src_dir: Path to the IDEA source directory.

    Returns:
        Sorted list of (filename, sheet_name) tuples from b_src_dir
        whose sheet_name contains a idea_type string.

    Raises:
        ValueError: If any idea_type sheet found in b_src_dir also exists
                    in i_src_dir (matched on sheet_name alone).
    """
    bele_ii_sheets = get_sheets_with_idea_types(b_src_dir)
    idea_ii_sheets = get_sheets_with_idea_types(i_src_dir)
    bele_ii_sheets_set = set(bele_ii_sheets)
    idea_ii_sheets_set = set(idea_ii_sheets)

    if overlapping := idea_ii_sheets_set.intersection(bele_ii_sheets_set):
        raise ValueError(
            f"idea_type sheets found in both b_src_dir and i_src_dir: "
            f"{sorted(overlapping)}"
        )

    return bele_ii_sheets


def beliefs_sheets_to_idea_sheets(
    b_src_dir: str, i_src_dir: str, db_max_spark_num: int = None
) -> List[Tuple[str, str]]:
    """
    Copies all idea_type sheets from b_src_dir into i_src_dir.
    Each idea_type sheet is written into its own new Excel file, named after the sheet,
    preserving values and structure for downstream pandas operations.

    Args:
        b_src_dir: Path to the BELE source directory.
        i_src_dir: Path to the IDEA source directory.

    Returns:
        Sorted list of (new_filename, sheet_name) tuples for every sheet copied.

    Raises:
        ValueError: (propagated from get_bele_ii_sheets_validated) if any BR
                    sheet name exists in both directories before the copy.
    """
    bele_spark_faces = get_spark_faces_from_files(b_src_dir)
    idea_max_spark_num = get_0_if_None(get_max_spark_num_from_files(i_src_dir))
    general_max_spark_num = max(idea_max_spark_num, get_0_if_None(db_max_spark_num))
    spark_face_spark_nums = create_spark_face_spark_nums(
        bele_spark_faces, general_max_spark_num
    )

    bele_ii_sheets = get_validated_b_src_idea_type_sheets(b_src_dir, i_src_dir)
    # Group sheet names by their source file
    file_to_sheets: dict[str, List[str]] = {}
    for filename, sheet_name in bele_ii_sheets:
        file_to_sheets.setdefault(filename, []).append(sheet_name)

    copied: List[Tuple[str, str]] = []

    for filename, sheet_names in file_to_sheets.items():
        src_path = os_path_join(b_src_dir, filename)
        dst_path = os_path_join(i_src_dir, filename)
        for sheet_name in sheet_names:
            ii_df = pandas_read_excel(src_path, sheet_name)
            add_spark_num_column(ii_df, spark_face_spark_nums)
            save_sheet(dst_path, sheet_name, ii_df, False)
            copied.append((dst_path, sheet_name))

    delete_dir(b_src_dir)
    set_dir(b_src_dir)
    return sorted(copied)
