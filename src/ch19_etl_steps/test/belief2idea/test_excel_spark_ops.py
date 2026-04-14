from openpyxl import Workbook as openpyxl_Workbook
from os.path import exists as os_path_exists, join as os_path_join
from pandas import (
    DataFrame,
    ExcelWriter as pandas_ExcelWriter,
    isna as pandas_isna,
    read_excel as pandas_read_excel,
)
from pathlib import Path
from pytest import fixture as pytest_fixture, raises as pytest_raises
from src.ch00_py.file_toolbox import count_dirs_files, create_path
from src.ch17_idea.idea_db_tool import sheet_exists
from src.ch19_etl_steps.belief2idea import (
    add_spark_num_column,
    beliefs_sheets_to_idea_sheets,
    create_spark_face_spark_nums,
    get_excel_sheet_tuples,
    get_max_spark_num_from_files,
    get_sheets_with_idea_types,
    get_spark_faces_from_df,
    get_spark_faces_from_files,
    get_validated_b_src_idea_type_sheets,
)
from src.ref.keywords import Ch19Keywords as kw, ExampleStrs as exx


def test_get_spark_faces_from_df_ReturnsObj_Scenario0_Basic():
    # ESTABLISH
    df = DataFrame({kw.spark_face: ["a", "b", "a", "c"]})
    # WHEN
    result = get_spark_faces_from_df(df)
    # THEN
    assert result == {"a", "b", "c"}


def test_get_spark_faces_from_df_ReturnsObj_Scenario1_excludes_nulls():
    # ESTABLISH
    df = DataFrame({kw.spark_face: ["a", None, "b", float("nan")]})
    # WHEN
    result = get_spark_faces_from_df(df)
    # THEN
    assert result == {"a", "b"}


def test_get_spark_faces_from_df_ReturnsObj_Scenario2_MissingColumnReturnsEmptySet():
    # ESTABLISH
    df = DataFrame({"other_col": [1, 2, 3]})
    # WHEN
    result = get_spark_faces_from_df(df)
    # THEN
    assert result == set()


def test_get_spark_faces_from_files_ReturnsObj_Scenario0_Multiple_files(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    file2 = tmp_path / "file2.xlsx"

    df1 = DataFrame({kw.spark_face: ["a", "b"]})
    df2 = DataFrame({kw.spark_face: ["b", "c"]})

    df1.to_excel(file1, index=False)
    df2.to_excel(file2, index=False)
    # WHEN
    result = get_spark_faces_from_files(tmp_path)
    # THEN
    assert result == {"a", "b", "c"}


def test_get_spark_faces_from_files_ReturnsObj_Scenario1_Multiple_sheets(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"

    df1 = DataFrame({kw.spark_face: ["a"]})
    df2 = DataFrame({kw.spark_face: [exx.sue]})

    with pandas_ExcelWriter(file1) as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
    # WHEN
    result = get_spark_faces_from_files(tmp_path)
    # THEN
    assert result == {"a", exx.sue}


def test_get_spark_faces_from_files_ReturnsObj_Scenario2_IgnoresMissingColumn(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"

    df1 = DataFrame({kw.spark_face: ["a"]})
    df2 = DataFrame({"other": [1, 2]})  # no spark_face column

    with pandas_ExcelWriter(file1) as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
    # WHEN
    result = get_spark_faces_from_files(tmp_path)
    # THEN
    assert result == {"a"}


def test_get_max_spark_num_from_files_ReturnsObj_Scenario0_MultipleFiles(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    file2 = tmp_path / "file2.xlsx"
    df1 = DataFrame({kw.spark_num: [1, 2, 3]})
    df2 = DataFrame({kw.spark_num: [4, 5]})
    df1.to_excel(file1, index=False)
    df2.to_excel(file2, index=False)
    # WHEN
    result = get_max_spark_num_from_files(tmp_path)
    # THEN
    assert result == 5


def test_get_max_spark_num_from_files_ReturnsObj_Scenario1_IgnoresInvalidAndConvertsFloats(
    tmp_path,
):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    df = DataFrame({kw.spark_num: ["10", "bad", None, 7.9]})  # 7.9 -> 7
    df.to_excel(file1, index=False)
    # WHEN
    result = get_max_spark_num_from_files(tmp_path)
    # THEN
    assert result == 10


def test_get_max_spark_num_from_files_ReturnsObj_Scenario2_MultipleSheetsAndMissingColumn(
    tmp_path,
):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    df1 = DataFrame({kw.spark_num: [1, 20]})
    df2 = DataFrame({"other": [100, 200]})  # no spark_num
    df3 = DataFrame({kw.spark_num: [15]})
    with pandas_ExcelWriter(file1) as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
        df3.to_excel(writer, sheet_name="Sheet3", index=False)

    # WHEN
    result = get_max_spark_num_from_files(tmp_path)
    # THEN
    assert result == 20


def test_create_spark_face_spark_nums_ReturnsObj_Scenario0_Simple():
    # ESTABLISH
    spark_faces = {exx.sue, exx.bob, exx.yao}
    max_spark_num = 11
    # WHEN
    x_dict = create_spark_face_spark_nums(spark_faces, max_spark_num)
    # THEN
    assert x_dict == {exx.bob: 12, exx.sue: 13, exx.yao: 14}


def test_create_spark_face_spark_nums_ReturnsObj_Scenario1_max_spark_num_IsNone():
    # ESTABLISH
    spark_faces = {exx.sue, exx.bob, exx.yao}
    max_spark_num = None
    # WHEN
    x_dict = create_spark_face_spark_nums(spark_faces, max_spark_num)
    # THEN
    assert x_dict == {exx.bob: 1, exx.sue: 2, exx.yao: 3}


def test_create_spark_face_spark_nums_ReturnsObj_Scenario2_No_max_spark_num():
    # ESTABLISH
    spark_faces = {exx.sue, exx.bob, exx.yao}
    # WHEN
    x_dict = create_spark_face_spark_nums(spark_faces)
    # THEN
    assert x_dict == {exx.bob: 1, exx.sue: 2, exx.yao: 3}


def test_add_spark_num_column_SetsAttr_Scenario0_Add_spark_num_Basic():
    # ESTABLISH
    df = DataFrame({kw.spark_face: ["a", "b", "c"]})
    mapping = {"a": 1, "b": 2, "c": 3}
    # WHEN
    add_spark_num_column(df, mapping)
    # THEN
    assert list(df.columns)[0] == kw.spark_num
    assert df[kw.spark_num].tolist() == [1, 2, 3]


def test_add_spark_num_column_SetsAttr_Scenario1_MissingSparkFaceSets_nan():
    # ESTABLISH
    df = DataFrame({kw.spark_face: ["a", "b", "x"]})
    mapping = {"a": 1, "b": 2}
    # WHEN
    add_spark_num_column(df, mapping)
    # THEN
    assert df[kw.spark_num].tolist()[:2] == [1, 2]
    assert pandas_isna(df[kw.spark_num].iloc[2])


def test_add_spark_num_column_SetsAttr_Scenario0_MutatesOriginalDataframe():
    # ESTABLISH
    df = DataFrame({kw.spark_face: ["a", "b"]})
    mapping = {"a": 1, "b": 2}
    assert kw.spark_num not in df.columns
    # WHEN
    add_spark_num_column(df, mapping)
    # THEN
    assert kw.spark_num in df.columns


@pytest_fixture
def excel_dir(tmp_path):
    """Creates a temporary directory with sample Excel files for testing."""

    # File 1: two sheets
    wb1 = openpyxl_Workbook()
    wb1.active.title = "Alpha"
    wb1.create_sheet("Beta")
    wb1.save(tmp_path / "report.xlsx")

    # File 2: one sheet
    wb2 = openpyxl_Workbook()
    wb2.active.title = "Summary"
    wb2.save(tmp_path / "data.xlsx")

    # Non-Excel file (should be ignored)
    (tmp_path / "notes.txt").write_text("ignore me")

    return tmp_path


def test_get_excel_sheet_tuples_ReturnsObj_Scenario0_AllSheetTuples(excel_dir):
    """All (filename, sheet) pairs across every Excel file are returned."""
    # ESTABLISH / WHEN
    result = get_excel_sheet_tuples(str(excel_dir))
    # THEN
    assert ("data.xlsx", "Summary") in result
    assert ("report.xlsx", "Alpha") in result
    assert ("report.xlsx", "Beta") in result
    assert len(result) == 3


def test_get_excel_sheet_tuples_ReturnsObj_Scenario1_SortedList(excel_dir):
    """Returned list is sorted lexicographically by (filename, sheet_name)."""
    # ESTABLISH / WHEN
    result = get_excel_sheet_tuples(str(excel_dir))
    # THEN
    assert result == sorted(result)


def test_get_excel_sheet_tuples_ReturnsObj_Scenario2_EmptyListForNoExcelFiles(
    tmp_path: Path,
):
    """Returns an empty list when the directory contains no Excel files."""
    # ESTABLISH
    (tmp_path / "readme.md").write_text("nothing here")
    # WHEN
    result = get_excel_sheet_tuples(str(tmp_path))
    # THEN
    assert result == []


def test_get_sheets_with_idea_types_ReturnsObj_Scenario0_MatchingTuples(
    tmp_path: Path,
):  # sourcery skip: extract-duplicate-method
    """Only tuples whose sheet_name contains a ii_string are returned."""
    # ESTABLISH
    beliefs_excel_dir = tmp_path / "beliefs"
    beliefs_excel_dir.mkdir()
    wb1 = openpyxl_Workbook()
    wb1.active.title = "ii00002_Sales"
    wb1.create_sheet("Revenue")
    wb1.create_sheet("Costs_ii00005")
    wb1.save(beliefs_excel_dir / "x300reports.xlsx")

    wb2 = openpyxl_Workbook()
    wb2.active.title = "Summary"
    wb2.create_sheet("ii00042_Overview")
    wb2.save(beliefs_excel_dir / "report.xlsx")

    # WHEN
    result = get_sheets_with_idea_types(beliefs_excel_dir)

    # THEN
    assert ("x300reports.xlsx", "ii00002_Sales") in result
    assert ("x300reports.xlsx", "Costs_ii00005") in result
    assert ("report.xlsx", "ii00042_Overview") in result
    assert ("x300reports.xlsx", "Revenue") not in result
    assert ("report.xlsx", "Summary") not in result


def test_get_validated_b_src_idea_type_sheets_ReturnsObj_Scenario0_BeleBrSheets(
    tmp_path,
):
    """Returns only idea_type sheet tuples from b_src_dir when there is no overlap."""
    # ESTABLISH
    bele_dir = tmp_path / "bele"
    bele_dir.mkdir()
    i_src_dir = tmp_path / "ideas"
    i_src_dir.mkdir()
    wb = openpyxl_Workbook()
    wb.active.title = "ii00005_Sales"
    wb.create_sheet("Revenue")
    wb.create_sheet("ii00042_Costs")
    wb.save(bele_dir / "x300reports.xlsx")

    # WHEN
    result = get_validated_b_src_idea_type_sheets(bele_dir, i_src_dir)
    # THEN
    assert ("x300reports.xlsx", "ii00005_Sales") in result
    assert ("x300reports.xlsx", "ii00042_Costs") in result
    assert ("x300reports.xlsx", "Revenue") not in result


def test_get_validated_b_src_idea_type_sheets_Scenario1_RaisesOnOverlap(
    tmp_path: Path,
):
    """Raises ValueError when a idea_type sheet name exists in both directories."""
    # ESTABLISH
    bele_dir = tmp_path / "bele"
    bele_dir.mkdir()
    bele_wb = openpyxl_Workbook()
    bele_wb.active.title = "ii00005_Sales"
    bele_wb.create_sheet("Revenue")
    bele_wb.create_sheet("ii00042_Costs")
    x3_filename = "x300reports.xlsx"
    bele_wb.save(bele_dir / x3_filename)

    i_src_dir = tmp_path / "idea_overlap"
    i_src_dir.mkdir()
    idea_wb = openpyxl_Workbook()
    idea_wb.active.title = "ii00005_Sales"  # overlaps with bele_dir
    idea_wb.save(i_src_dir / x3_filename)

    # WHEN / THEN
    with pytest_raises(ValueError, match="ii00005_Sales"):
        get_validated_b_src_idea_type_sheets(bele_dir, i_src_dir)


def test_get_validated_b_src_idea_type_sheets_Scenario2_DoesNotRaiseError(
    tmp_path: Path,
):
    """Raises ValueError when a idea_type sheet name exists in both directories."""
    # ESTABLISH
    bele_dir = tmp_path / "bele"
    bele_dir.mkdir()
    bele_wb = openpyxl_Workbook()
    bele_wb.active.title = "ii00005_Sales"
    bele_wb.create_sheet("Revenue")
    ii42_sheetname = "ii00042_Costs"
    bele_wb.create_sheet(ii42_sheetname)
    x3_filename = "x300reports.xlsx"
    bele_wb.save(bele_dir / x3_filename)

    i_src_dir = tmp_path / "idea_overlap"
    i_src_dir.mkdir()
    idea_wb = openpyxl_Workbook()
    idea_wb.active.title = "ii00005_Sales"  # overlaps with bele_dir
    x4_filename = "x400reports.xlsx"
    idea_wb.create_sheet(ii42_sheetname)
    idea_wb.save(i_src_dir / x4_filename)

    # WHEN
    sheet_tuples = get_validated_b_src_idea_type_sheets(bele_dir, i_src_dir)
    # THEN
    print(f"{(x3_filename, ii42_sheetname)=}")
    print(f"{sheet_tuples=}")
    assert (x3_filename, ii42_sheetname) in sheet_tuples


def test_get_validated_b_src_idea_type_sheets_ReturnsObj_Scenario2_EmptyWhenNoBeleBrSheets(
    tmp_path,
):
    """Returns an empty list when b_src_dir has no idea_type sheets."""
    # ESTABLISH
    i_src_dir = tmp_path / "ideas"
    i_src_dir.mkdir()
    wb = openpyxl_Workbook()
    wb.active.title = "Summary"
    wb.create_sheet("Details")
    wb.save(i_src_dir / "idea_report.xlsx")

    empty_bele = tmp_path / "empty_bele"
    empty_bele.mkdir()
    wb = openpyxl_Workbook()
    wb.active.title = "Summary"
    wb.save(empty_bele / "plain.xlsx")
    # WHEN
    result = get_validated_b_src_idea_type_sheets(empty_bele, i_src_dir)
    # THEN
    assert result == []


def test_beliefs_sheets_to_idea_sheets_Scenario0_TwoTuples(tmp_path: Path):
    """Returns one (filename, sheet_name) tuple per idea_type sheet copied."""
    # ESTABLISH
    empty_i_src_dir = tmp_path / "ideas"
    empty_i_src_dir.mkdir()

    populated_bele_dir = tmp_path / "bele"
    populated_bele_dir.mkdir()
    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00020_Sales"
    ws1.append(["product", "units", "revenue"])
    ws1.append(["widget", 10, 500])
    ws1.append(["gadget", 5, 250])

    ws2 = wb.create_sheet("ii00004_Costs")
    ws2.append(["category", "amount"])
    ws2.append(["rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_bele_dir / "AllSales.xlsx")

    # WHEN
    result = beliefs_sheets_to_idea_sheets(populated_bele_dir, empty_i_src_dir)
    # THEN
    dst_all_sales_path = create_path(empty_i_src_dir, "AllSales.xlsx")
    assert (dst_all_sales_path, "ii00004_Costs") in result
    assert (dst_all_sales_path, "ii00020_Sales") in result
    assert len(result) == 2


def test_beliefs_sheets_to_idea_sheets_Scenario1_CreatesDestinationFile(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    empty_i_src_dir = tmp_path / "ideas"
    empty_i_src_dir.mkdir()
    populated_bele_dir = tmp_path / "bele"
    populated_bele_dir.mkdir()
    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00020_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    ws1.append([exx.sue, "gadget", 5, 250])

    ws2 = wb.create_sheet("ii00004_Costs")
    ws2.append([kw.spark_face, "category", "amount"])
    ws2.append([exx.sue, "rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_bele_dir / "AllSales.xlsx")

    # WHEN
    beliefs_sheets_to_idea_sheets(populated_bele_dir, empty_i_src_dir)
    # THEN
    allsales_path = os_path_join(str(empty_i_src_dir), "AllSales.xlsx")
    df = pandas_read_excel(allsales_path, sheet_name="ii00020_Sales")
    expected_dst_columns = [kw.spark_num, kw.spark_face, "product", "units", "revenue"]
    assert list(df.columns) == expected_dst_columns
    assert len(df) == 2
    assert df[kw.spark_num].min() == 1
    assert df["revenue"].sum() == 750


def test_beliefs_sheets_to_idea_sheets_Scenario2_RaisesOnOverlap(tmp_path: Path):
    # sourcery skip: extract-duplicate-method
    """Propagates ValueError from get_bele_ii_sheets_validated on sheet name overlap."""
    # ESTABLISH
    beliefs_dir = tmp_path / "bele"
    beliefs_dir.mkdir()
    i_src_dir = tmp_path / "ideas"
    i_src_dir.mkdir()

    wb_bele = openpyxl_Workbook()
    wb_bele.active.title = "ii00020_Sales"
    allsales_filename = "AllSales.xlsx"
    wb_bele.save(beliefs_dir / allsales_filename)

    wb_idea = openpyxl_Workbook()
    wb_idea.active.title = "ii00020_Sales"
    wb_idea.save(i_src_dir / allsales_filename)
    # WHEN / THEN
    with pytest_raises(ValueError, match="ii00020_Sales"):
        beliefs_sheets_to_idea_sheets(beliefs_dir, i_src_dir)


def test_beliefs_sheets_to_idea_sheets_Scenario3_DestinationFileHas_spark_num_SetBy_i_src_dir(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    i_src_dir = tmp_path / "ideas"
    i_src_dir.mkdir()
    populated_bele_dir = tmp_path / "bele"
    populated_bele_dir.mkdir()
    idea_wb = openpyxl_Workbook()
    idea_ws1 = idea_wb.active
    idea_ws1.title = "ii00020_Sales"
    expected_dst_columns = [kw.spark_num, kw.spark_face, "product", "units", "revenue"]
    idea_ws1.append(expected_dst_columns)
    curr_spark_num = 10
    idea_ws1.append([curr_spark_num, exx.sue, "widget", 10, 500])
    idea_wb.save(i_src_dir / "OtherFile.xlsx")

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00020_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    ws1.append([exx.sue, "gadget", 5, 250])

    ws2 = wb.create_sheet("ii00004_Costs")
    ws2.append([kw.spark_face, "category", "amount"])
    ws2.append([exx.sue, "rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_bele_dir / "AllSales.xlsx")

    # WHEN
    beliefs_sheets_to_idea_sheets(populated_bele_dir, i_src_dir)
    # THEN
    allsales_path = os_path_join(str(i_src_dir), "AllSales.xlsx")
    df = pandas_read_excel(allsales_path, sheet_name="ii00020_Sales")
    assert list(df.columns) == expected_dst_columns
    assert len(df) == 2
    assert df[kw.spark_num].min() == 11
    assert df[kw.spark_num].min() == curr_spark_num + 1
    assert df["revenue"].sum() == 750


def test_beliefs_sheets_to_idea_sheets_Scenario4_ParameterSparkNumAccepted(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    i_src_dir = tmp_path / "ideas"
    i_src_dir.mkdir()
    populated_bele_dir = tmp_path / "bele"
    populated_bele_dir.mkdir()
    idea_wb = openpyxl_Workbook()
    idea_ws1 = idea_wb.active
    idea_ws1.title = "ii00020_Sales"
    expected_dst_columns = [kw.spark_num, kw.spark_face, "product", "units", "revenue"]
    idea_ws1.append(expected_dst_columns)
    curr_spark_num = 10
    idea_ws1.append([curr_spark_num, exx.sue, "widget", 10, 500])
    idea_wb.save(i_src_dir / "OtherFile.xlsx")

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00020_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    ws1.append([exx.sue, "gadget", 5, 250])

    ws2 = wb.create_sheet("ii00004_Costs")
    ws2.append([kw.spark_face, "category", "amount"])
    ws2.append([exx.sue, "rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_bele_dir / "AllSales.xlsx")
    db_max_spark_num = 22

    # WHEN
    beliefs_sheets_to_idea_sheets(populated_bele_dir, i_src_dir, db_max_spark_num)
    # THEN
    allsales_path = os_path_join(str(i_src_dir), "AllSales.xlsx")
    df = pandas_read_excel(allsales_path, sheet_name="ii00020_Sales")
    assert df[kw.spark_num].min() != 11
    assert df[kw.spark_num].min() != curr_spark_num + 1
    assert df[kw.spark_num].min() == db_max_spark_num + 1


def test_beliefs_sheets_to_idea_sheets_Scenario5_src_dir_IsEmptied(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    bele_dir = tmp_path / "bele"
    bele_dir.mkdir()
    i_src_dir = tmp_path / "ideas"
    i_src_dir.mkdir()

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00020_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    wb.save(bele_dir / "AllSales.xlsx")
    assert count_dirs_files(bele_dir) == 1

    # WHEN
    beliefs_sheets_to_idea_sheets(bele_dir, i_src_dir)
    # THEN
    assert count_dirs_files(bele_dir) == 0


def test_beliefs_sheets_to_idea_sheets_Scenario6_src_num_Exists(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    bele_dir = tmp_path / "bele"
    bele_dir.mkdir()
    i_src_dir = tmp_path / "ideas"
    i_src_dir.mkdir()

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00020_Sales"
    ws1.append([kw.spark_num, kw.spark_face, "product", "units", "revenue"])
    ws1.append(["", exx.sue, "widget", 10, 500])
    bele_allsales_path = bele_dir / "AllSales.xlsx"
    wb.save(bele_allsales_path)
    assert count_dirs_files(bele_dir) == 1
    assert count_dirs_files(i_src_dir) == 0

    # WHEN
    beliefs_sheets_to_idea_sheets(bele_dir, i_src_dir)
    # THEN
    assert count_dirs_files(bele_dir) == 0
    assert count_dirs_files(i_src_dir) == 1
    idea_allsales_path = i_src_dir / "AllSales.xlsx"
    df = pandas_read_excel(idea_allsales_path, sheet_name="ii00020_Sales")
    assert df[kw.spark_num].min() == 1
